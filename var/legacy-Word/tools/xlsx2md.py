"""xlsx -> md 转换脚本（每个工作表渲染为一张 md 表格）

用法:
  python xlsx2md.py input.xlsx [output.md]

依赖: openpyxl
"""
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print('ERROR: 需要 openpyxl，请运行: pip install openpyxl')
    sys.exit(1)


def cell_text(v):
    if v is None:
        return ''
    return str(v).replace('\n', ' ').replace('|', '\\|').strip()


def convert(xlsx_path, md_path=None):
    if md_path is None:
        md_path = os.path.splitext(xlsx_path)[0] + '.md'

    if not os.path.exists(xlsx_path):
        print(f'ERROR: File not found: {xlsx_path}')
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)
    lines = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 去掉尾部全空行
        while rows and all(c is None or str(c).strip() == '' for c in rows[-1]):
            rows.pop()
        if not rows:
            continue

        lines.append(f'## {ws.title}')
        lines.append('')

        # 表头
        ncols = max(len(r) for r in rows)
        header = rows[0] + (None,) * (ncols - len(rows[0]))
        lines.append('| ' + ' | '.join(cell_text(c) for c in header) + ' |')
        lines.append('| ' + ' | '.join('---' for _ in header) + ' |')
        for r in rows[1:]:
            r = r + (None,) * (ncols - len(r))
            cells = [cell_text(c) for c in r]
            # 跳过全空行
            if all(c == '' for c in cells):
                continue
            lines.append('| ' + ' | '.join(cells) + ' |')
        lines.append('')

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')

    print(f'DONE: {md_path}')
    print(f'  Sheets: {len(wb.worksheets)}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    xlsx_file = sys.argv[1]
    md_file = sys.argv[2] if len(sys.argv) > 2 else None
    convert(xlsx_file, md_file)
